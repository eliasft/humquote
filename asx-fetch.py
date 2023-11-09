# -*- coding: utf-8 -*-
"""
Spyder Editor

Script file for fetching ASX electricity futures.
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
import sqlite3
from datetime import datetime

from openpyxl import load_workbook

# Function to save dataframe to an Excel file
def save_to_excel(df, filename):
    # Extract the year from the quote date for the sheet name
    sheet_name = str(df['quote_date'].iloc[0].year)

    # Check if the Excel file already exists
    try:
        # Attempt to load the workbook
        book = load_workbook(filename)
        # If the sheet exists, append data, otherwise create a new sheet
        if sheet_name in book.sheetnames:
            with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
                writer.book = book
                writer.sheets = {ws.title: ws for ws in book.worksheets}
                
                # Get the last row in the existing sheet
                # Assuming the sheet is not empty and has a header
                startrow = writer.sheets[sheet_name].max_row
                
                # Append data without adding the header
                df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False, header=False)
        else:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                writer.book = book
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        # If the workbook does not exist, create it
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


import requests
from bs4 import BeautifulSoup
import pandas as pd
import sqlite3
from datetime import datetime

# The target URL
url = 'https://www.asxenergy.com.au'

# Use the 'requests' library to perform an HTTP GET request
response = requests.get(url)

# Check if the request was successful
if response.status_code == 200:
    # Use BeautifulSoup to parse the HTML content
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Find the date of the quotes
    date_tag = soup.find('h3', string=lambda t: t and 'Cal Base Future Prices' in t)
    if date_tag:
        # Extract the date string and convert it to a datetime object
        date_str = date_tag.get_text().replace('Cal Base Future Prices ', '')
        quote_date = datetime.strptime(date_str, '%a %d %b %Y').date()
    else:
        quote_date = datetime.now().date()  # Fallback to current date if not found
    
    # Find the futures prices table by its unique attributes or structure
    prices_table = soup.find('div', class_='dataset')
    
    # Check if we found the table
    if prices_table:
        # Extract the rows from the table
        rows = prices_table.find_all('tr')
        #print(rows)
        # Initialize an empty list to store the data
        data = []
        
        # Iterate over the rows to extract the data
        for row in rows[1:]:  # Skip the header row
            # Get all cells in the row
            cells = row.find_all('td')
            #print('Cells',cells)
            # Extract the text from each cell and add it to the row data
            year_of_instrument = cells[0].get_text().strip()
            #print('year of instrument',year_of_instrument)
            # Handle whitespace or non-numeric characters in the year column
            year_of_instrument = ''.join(filter(str.isdigit, year_of_instrument))
            row_data = [quote_date, int(year_of_instrument)] + [cell.get_text().strip() for cell in cells[1:]]
            data.append(row_data)
        
        # Define the headers separately to avoid whitespace issues
        headers = ['quote_date', 'instrument_year', 'NSW', 'VIC', 'QLD', 'SA']
        
        # Create a pandas DataFrame from the data
        df = pd.DataFrame(data, columns=headers)
        
        # Clean up the DataFrame - convert numeric columns to floats
        for col in df.columns[2:]:  # Skip the first two columns (date and instrument_year)
            df[col] = pd.to_numeric(df[col], errors='coerce')
            
        # Call the function to save the DataFrame to an Excel file
        save_to_excel(df, 'futures_prices_database.xlsx')
    
        display(df)
    else:
        print("The futures prices table was not found.")
else:
    print(f"Failed to retrieve the page. Status code: {response.status_code}")
